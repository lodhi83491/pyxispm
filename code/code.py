import pandas as pd
from os import walk
from openpyxl import load_workbook
import formulas
from pandas import ExcelWriter
import numpy as np
import time


#Path of input files

mypath = r"C:\Users\phoolsinghlodhi\Desktop\pyxispm"

#number of rows per chunk
chunk_size=1000000
all_formula ={}


excel_col = {"A":"CUSTOMER_NAME","B":"HNW_TYPE","C":"HNW_CATEGORY","D":"AL_TAG","E":"AL_AMT","F":"PL_TAG","G":"PL_AMT","H":"BL_TAG","I":"BL_AMT","J":"TWL_AMT","K":"INSTA_LOAN_AMT","L":"INSTA_LOAN_CARD_NO","M":"JUMBO_LOAN_AMT","N":"JUMBO_LOAN_CARD_NO","O":"DC_EMI_AMT","P":"LAST_4_DIGITS_DC_EMI","Q":"CC_PRE_APPROVED_TAG","R":"CC_PRE_APPROVED_CARD","S":"LAST_4_DIGITS_FOR_DC","T":"LAST_4_DIGITS_FOR_CC","U":"BRANCH_CODE","V":"GENDER"}

def NumberOfInputFiles(mypath):
    print("NumberOfInputFiles")
    ListOfFiles = []
    for (dirpath, dirnames, filenames) in walk(mypath+"\\data"):
        ListOfFiles.extend(filenames)
        break
    return ListOfFiles


def formula(mypath):
    print("Inside Formula")
    wb = load_workbook(mypath+"\\template\\WORKING-tmp.xlsx")
    formula = []
    columns =[i for i in range(1,50)]
    sheet = wb["Sheet1"]
    for ii in range(1,50):
        if str((sheet.cell(row=3,column=ii).value)).strip()[0:1]!="=":
            columns.remove(ii)
        else:
            formula.append(sheet.cell(row=3,column=ii).value)
    all_formula["columns"]=columns
    all_formula["formula"]=formula
    
   
def Processing(data,pre):
    start_time = time.time()
    data.replace(np.NaN, '', inplace=True)
    print("Inside Processing")
    length =len(data)
    headers=data.columns
    output={"id":[], "name":[], "al_link":[],"al_showslide2":[],"al_pre-approvedText":[],"al_loanAmount":[],"twl_showslide2":[],"twl_loanAmount":[],"twl_link":[],"pl_loanType":[],"pl_preapprovedText":[],"pl_loanAmount":[],"pl_link":[],"hl_name":[],"hl_link":[],"instaloan_show":[],"instaloan_link":[],"instaloan_amount":[],"instaloan_cardno":[],"consumerLoan_showslide":[],"consumerloan_amount":[],"consumerloan_cardno":[],"consumerloan_link":[],"cc_cardno":[],"cc_link":[],"url":[]}
    ids=pd.read_excel(mypath+"\\random id\\random IDs - 10 Lac.xlsx",header=None)
    output["id"].extend([pre+str(i).lower() for i in ids[0].iloc[:length,]])
    #print("Id :",output["id"])
    output["twl_link"].extend(["https://leads.hdfcbank.com/applications/new_webforms/apply/mobile/Two-wheeler-Loan.aspx?pcode=QLS_VIDEO_VIS"]*length)
    output["hl_name"].extend(['']*length)
    output["hl_link"].extend(["https://leads.hdfcbank.com/applications/new_webforms/apply/mobile/Home-Loan-Indians.aspx?SourceCode=paid&pcode=QLS_EMAIL&promocode=QLS_EMAIL&utm_campaign=QLS&utm_medium=email&utm_source=QLS_EMAIL"]*length)
    output["instaloan_link"].extend(["https://leads.hdfcbank.com/applications/webforms/apply/CC_term_loan_jumbo_loan/default.aspx?sourcecode=FT_microsite_jumbo&pcode=QLS_EMAIL&promocode=QLS_EMAIL&utm_campaign=QLS&utm_medium=email&utm_source=QLS_EMAIL"]*length)
    output["consumerloan_link"].extend(["https://v1.hdfcbank.com/htdocs/common/festive-treats/loan_product/easyemi.html?pcode=QLS_EMAIL&promocode=QLS_EMAIL&utm_campaign=%20QLS&utm_medium=email&utm_source=QLS_EMAIL"]*length)
    output["url"].extend(["https://mf1.pctr.co/"+str(i) for i in output["id"]])
    count_final=0
    count=0
    counter=1
    print("Parser function ")
    for i in all_formula["formula"][:-1]:
        count+=1
        one_col=[]
        func=formulas.Parser().ast(str(i))[1].compile()
        col=list(func.inputs)[0][:-1]
        m=excel_col[col]
        try:
            for c in data[m]:
                one_col.append(func(c))
            if counter ==2:
                output["al_link"].extend(one_col)
            elif counter==3:
                output["al_showslide2"].extend(one_col)
            elif counter==4:
                output["al_pre-approvedText"].extend(one_col)
            elif counter==6:
                output["twl_showslide2"].extend(one_col)
            elif counter==8:
                output["pl_loanType"].extend(one_col)
            elif counter==9:
                output["pl_preapprovedText"].extend(one_col)
            elif counter==11:
                output["pl_link"].extend(one_col)
            elif counter==12:
                output["instaloan_show"].extend(one_col)
            elif counter==15:
                output["consumerLoan_showslide"].extend(one_col)
            elif counter==19:
                output["cc_link"].extend(one_col)
        except:
            if count==1 and counter==1:
                #for iii in data["CUSTOMER_NAME"]:
                    #if len(iii)!=0:
                        #output["name"].append(iii.split()[0])
                    #else:
                        #output["name"].append("")
                output["name"].extend([i.split()[0] if len(i)!=0 else "" for i in data["CUSTOMER_NAME"]])
            else:
                if count in [5,7,10,13,16]:
                    for c in data[m]:
                        flag=False
                        cc=""
                        if c:
                            flag=True
                            s=""
                            cc=str(c).split(".")[0]
                            l =len(cc)
                            count1=0
                            while(l>0):
                                l-=1
                                if count1 in [3,5,7,9,11,13,15]:
                                    s+=","
                                s+=cc[l]
                                count1+=1
                            cc=s[::-1]
                        if counter==5:
                            if flag:
                                output["al_loanAmount"].append(cc)
                            else:
                                output["al_loanAmount"].append('')
                        elif counter == 7:
                            if flag:
                                output["twl_loanAmount"].append(cc)
                            else:
                                output["twl_loanAmount"].append('')
                        elif counter == 10:
                            if flag:
                                output["pl_loanAmount"].append(cc)
                            else:
                                output["pl_loanAmount"].append('')
                        elif counter == 13:
                            if flag:
                                output["instaloan_amount"].append(cc)
                            else:
                                output["instaloan_amount"].append('')
                        elif counter == 16:
                            if flag:
                                output["consumerloan_amount"].append(cc)
                            else:
                                output["consumerloan_amount"].append('')
                
                elif count in [14,17,18]:
                    for c in data[m]:
                        flag=False
                        cc=""
                        if c:
                            flag=True
                            cc="xx"+str(c).split(".")[0]
                        if counter==14:
                            if flag:
                                output["instaloan_cardno"].append(cc)
                            else:
                                output["instaloan_cardno"].append('')
                        elif counter == 17:
                            if flag:
                                output["consumerloan_cardno"].append(cc)
                            else:
                                output["consumerloan_cardno"].append('')
                        elif counter == 18:
                            if flag:
                                output["cc_cardno"].append(cc)
                            else:
                                output["cc_cardno"].append('')

        counter+=1
        print("Counter : ",counter)
    for k,v in output.items():
        print(k,len(v))
    #print("+++++++++++++++++++++")
    df = pd.DataFrame(output)
    print("Total Processing time--- %s seconds ---" % (time.time() - start_time))
    return df



def CsvFile(fileName,num):
    print("CSV file :",fileName)
    pre = ['a','b','c','d']
    c=0
    n=0
    #c_columns=["CUSTOMER_NAME","HNW_TYPE","HNW_CATEGORY","AL_TAG","AL_AMT","PL_TAG","PL_AMT","BL_TAG","BL_AMT","TWL_AMT","INSTA_LOAN_AMT","INSTA_LOAN_CARD_NO","JUMBO_LOAN_AMT","JUMBO_LOAN_CARD_NO","DC_EMI_AMT","LAST_4_DIGITS_DC_EMI","CC_PRE_APPROVED_TAG","CC_PRE_APPROVED_CARD","LAST_4_DIGITS_FOR_DC","LAST_4_DIGITS_FOR_CC","BRANCH_CODE","GENDER"]
    #for chunk in pd.read_csv(mypath+"\\data\\"+fileName, chunksize=chunk_size,skiprows=1, names=c_columns, header=None):
    for chunk in pd.read_csv(mypath+"\\data\\"+fileName, chunksize=chunk_size):
        n+=1
        print("Chunk length :",len(chunk))
        df = Processing(chunk,pre[c])
        writer = ExcelWriter(mypath+"\\output\\output"+str(num)+""+str(n)+".xlsx", options={"strings_to_urls":False})
        df.to_excel(writer,'Sheet1', index=False)
        writer.save()
        c+=1
        print("N :",n)
        print("Num :",num)
        if c==4:
            c=0
            
          
def ExcelFile(fileName,num):
    print("Excel file :",fileName)
    data = pd.read_excel(mypath+"\\data\\"+fileName)
    
    
def TextFile(fileName,num):
    print("Text File :",fileName)
    f = open(mypath+"\\data\\"+fileName)
    c=0
    while 1:
        lines = f.readlines(1000000)
        if not lines:
            break
        else:
            c+=1
                      
            
def CheckFileType(ListOfFiles):
    print("CheckFileType")
    num=0
    for files in ListOfFiles:
        num+=1
        file_type=files.split(".")
        if file_type[-1].lower() == "csv":
            CsvFile(files,num)
        elif file_type[-1].lower() == "xlsx":
            ExcelFile(files,num)
        elif file_type[-1].lower() == "txt":
            TextFile(files,num)
        else:
            print("Please Pass the data either in CSVor Excel or Text files...Thank you.....!!!")
            
            
if __name__=="__main__":
    formula(mypath)
    print("*******************************")
    files=NumberOfInputFiles(mypath)
    CheckFileType(files)

